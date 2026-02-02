import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from datetime import datetime
import os

class LiveEvaluationTracker:
    def __init__(self, excel_file: str = None):
        if excel_file:
            self.excel_file = excel_file
        else:
            self.excel_file = f"rag_live_evaluation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Initialize Excel file
        self._init_excel()
        
        self.questions = []
    
    def _init_excel(self):
        """Initialize Excel file with headers and formatting"""
        wb = Workbook()
        ws = wb.active
        ws.title = "RAG Evaluation"
        
        # Headers
        headers = [
            "Timestamp", "Question", "Answer", "Expected", "Correct",
            "Failure Type", "Retrieved Chunks", "Chunk Samples",
            "Query Complexity", "Top K", "Chunk Scores", "Chunk Size",
            "Response Time", "Document Type", "File Size"
        ]
        
        ws.append(headers)
        
        # Format headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(self.excel_file)
    
    def add_question(self, question_data: dict):
        """Add a new question to the tracker"""
        self.questions.append(question_data)
        self._update_excel(question_data)
    
    def _update_excel(self, data: dict):
        """Update Excel file with new data"""
        wb = openpyxl.load_workbook(self.excel_file)
        ws = wb.active
        
        # Prepare row data
        row = [
            data.get('timestamp', datetime.now().isoformat()),
            data.get('question', ''),
            data.get('answer', ''),
            data.get('expected', ''),
            "YES" if data.get('is_correct') else "NO",
            data.get('failure_type', 'NONE'),
            data.get('retrieved_chunks_count', 0),
            str(data.get('chunk_samples', []))[:200],
            data.get('debug_info', {}).get('query_complexity', ''),
            data.get('debug_info', {}).get('top_k', ''),
            str(data.get('debug_info', {}).get('chunk_scores', []))[:100],
            data.get('debug_info', {}).get('chunk_size_used', ''),
            data.get('response_time', ''),
            data.get('document_type', ''),
            data.get('file_size', '')
        ]
        
        ws.append(row)
        
        # Color code based on correctness
        row_num = ws.max_row
        correct_cell = ws.cell(row=row_num, column=5)
        
        if data.get('is_correct'):
            correct_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            correct_cell.font = Font(color="006100")
        else:
            correct_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            correct_cell.font = Font(color="9C0006")
        
        # Color failure types
        failure_cell = ws.cell(row=row_num, column=6)
        failure_type = data.get('failure_type', '').lower()
        
        if 'retrieval' in failure_type:
            failure_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        elif 'hallucination' in failure_type:
            failure_cell.fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
        
        wb.save(self.excel_file)
        
        print(f"✅ Question logged to Excel: {self.excel_file}")
    
    def get_summary_stats(self) -> dict:
        """Get summary statistics"""
        if not self.questions:
            return {}
        
        df = pd.DataFrame(self.questions)
        
        stats = {
            'total_questions': len(df),
            'correct_answers': df['is_correct'].sum(),
            'accuracy': df['is_correct'].mean() * 100,
            'retrieval_success_rate': (df['retrieved_chunks_count'] > 0).mean() * 100,
            'common_failure_types': df['failure_type'].value_counts().to_dict(),
            'avg_chunks_retrieved': df['retrieved_chunks_count'].mean()
        }
        
        return stats
    
    def save_report(self):
        """Save comprehensive report"""
        stats = self.get_summary_stats()
        
        with pd.ExcelWriter(self.excel_file, engine='openpyxl', mode='a') as writer:
            # Create summary sheet
            summary_df = pd.DataFrame([stats])
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Create charts sheet
            charts_data = {
                'Metric': ['Accuracy', 'Retrieval Success', 'Avg Chunks'],
                'Value': [stats['accuracy'], stats['retrieval_success_rate'], stats['avg_chunks_retrieved']]
            }
            charts_df = pd.DataFrame(charts_data)
            charts_df.to_excel(writer, sheet_name='Charts', index=False)
        
        print(f"📊 Report saved: {self.excel_file}")
        return self.excel_file